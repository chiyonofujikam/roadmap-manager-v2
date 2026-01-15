"""API routes for the application"""

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Body, File, HTTPException, UploadFile, status
from openpyxl import load_workbook

from rm_be.api.deps import (CurrentUser, RequireAdminOrResponsible,
                            RequireCollaborator)
from rm_be.api.schemas import (ActiveLCUpdate, ConditionalListCreate,
                               LCItemCreate, LCItemUpdate, LCMergeRequest,
                               ModificationRequestCreate,
                               ModificationRequestReview, PointageEntryCreate,
                               PointageEntryUpdate, UserCreate, UserUpdate)
from rm_be.api.utils import (get_active_lc_name, get_cstr_semaine,
                             get_db_user_from_current, serialize_date,
                             set_active_lc_name)
from rm_be.database import (ConditionalList, ConditionalListItem,
                            ConditionalListRepository, ModificationRequest,
                            ModificationRequestRepository, PointageEntry,
                            PointageEntryData, PointageEntryRepository, User,
                            UserRepository)

router = APIRouter(prefix="/api/v1", tags=["api"])


@router.get("/conditional-lists/default/items")
async def get_default_lc_items(current_user: dict = CurrentUser):
    """
    Get active items from the active LC (Liste Conditionnelle).
    This endpoint is accessible to all authenticated users (collaborators, responsibles, admins).
    Returns the LC items formatted for frontend autocomplete components.
    """
    def format_options(values):
        """Format values as options for AutocompleteInput component"""
        return [
            {
                "id": str(i),
                "label": value,
                "value": value,
                "active": True
            }
            for i, value in enumerate(sorted(values), 1)
        ]

    try:
        repo = ConditionalListRepository()
        active_lc_name = await get_active_lc_name()
        active_lc = await repo.find_by_name(active_lc_name)
        if not active_lc:
            return {
                "clef_imputation": [],
                "libelle": [],
                "fonction": []
            }

        active_items = await repo.find_active_items(active_lc["_id"])
        if not active_items:
            return {
                "clef_imputation": [],
                "libelle": [],
                "fonction": []
            }

        clef_imputation_set = set()
        libelle_set = set()
        fonction_set = set()
        for item in active_items:
            if item.get("clef_imputation"):
                clef_imputation_set.add(item["clef_imputation"])

            if item.get("libelle"):
                libelle_set.add(item["libelle"])

            if item.get("fonction"):
                fonction_set.add(item["fonction"])

        return {
            "clef_imputation": format_options(clef_imputation_set),
            "libelle": format_options(libelle_set),
            "fonction": format_options(fonction_set)
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching LC items: {str(e)}"
        )

@router.get("/pointage/team-entries")
async def get_team_pointage_entries(
    current_user: dict = RequireAdminOrResponsible(), 
    skip: int = 0, 
    limit: int = 1000,
    week_start: Optional[str] = None):
    """
    Get all pointage entries for a responsible's team.

    For responsible users: Returns entries for their team members only.
    For admin users: Returns entries for all users.
    Returns entries with user information included for table display.

    Args:
        current_user: Authenticated user (admin or responsible)
        skip: Number of entries to skip for pagination
        limit: Maximum number of entries to return
        week_start: Optional week start date (YYYY-MM-DD) to filter entries by week

    Returns:
        Dictionary with entries, total count, skip, and limit
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)

        user_type = db_user.get("user_type", current_user.get("user_type", ""))
        responsible_id = db_user.get("_id")

        query = {"is_deleted": {"$ne": True}}
        if week_start:
            try:
                week_start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
                cstr_semaine = get_cstr_semaine(week_start_date)
                query["entry_data.cstr_semaine"] = cstr_semaine
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid week_start format. Use YYYY-MM-DD"
                )

        if user_type == "admin":
            entries = await pointage_repo.find_many(
                query,
                skip=skip,
                limit=limit,
                sort=[("entry_data.date_pointage", -1), ("created_at", -1)]
            )
        else:
            team_entries = await pointage_repo.find_by_team(
                responsible_id,
                skip=skip,
                limit=limit
            )
            if week_start:
                cstr_semaine = query.get("entry_data.cstr_semaine")
                entries = [e for e in team_entries if e.get("entry_data", {}).get("cstr_semaine") == cstr_semaine]
            else:
                entries = team_entries
        user_ids = list(set([entry.get("user_id") for entry in entries if entry.get("user_id")]))
        users_dict = {}

        for user_id in user_ids:
            try:
                if isinstance(user_id, str) and ObjectId.is_valid(user_id):
                    user_obj = await user_repo.find_by_id(ObjectId(user_id))

                elif isinstance(user_id, ObjectId):
                    user_obj = await user_repo.find_by_id(user_id)

                else:
                    continue

                if user_obj:
                    users_dict[str(user_id)] = {
                        "id": str(user_obj.get("_id")),
                        "name": user_obj.get("name", "Unknown"),
                        "email": user_obj.get("email", ""),
                    }

            except Exception:
                continue

        formatted_entries = []
        for entry in entries:
            user_id = str(entry.get("user_id", ""))
            user_info = users_dict.get(user_id, {"name": "Unknown", "email": ""})
            entry_data = entry.get("entry_data", {})
            date_pointage_str = serialize_date(entry_data.get("date_pointage"))
            date_besoin_str = serialize_date(entry_data.get("date_besoin"))

            formatted_entries.append({
                "id": str(entry.get("_id", "")),
                "user_id": user_id,
                "user_name": user_info.get("name", "Unknown"),
                "user_email": user_info.get("email", ""),
                "date_pointage": date_pointage_str,
                "cstr_semaine": entry_data.get("cstr_semaine"),
                "clef_imputation": entry_data.get("clef_imputation", ""),
                "libelle": entry_data.get("libelle", ""),
                "fonction": entry_data.get("fonction", ""),
                "date_besoin": date_besoin_str,
                "heures_theoriques": entry_data.get("heures_theoriques", ""),
                "heures_passees": entry_data.get("heures_passees", ""),
                "commentaires": entry_data.get("commentaires", ""),
                "status": entry.get("status", "draft"),
                "created_at": entry.get("created_at"),
                "updated_at": entry.get("updated_at"),
                "submitted_at": entry.get("submitted_at"),
                "validated_at": entry.get("validated_at"),
            })

        return {
            "entries": formatted_entries,
            "total": len(formatted_entries),
            "skip": skip,
            "limit": limit
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching team pointage entries: {str(err)}"
        )

@router.get("/users/team-members")
async def get_team_members(current_user: dict = RequireAdminOrResponsible()):
    """
    Get all team members for a responsible user.

    For responsible users: Returns their team members (collaborators).
    For admin users: Returns all collaborators.

    Returns:
        List of user dictionaries with id, name, and email
    """
    try:
        user_repo = UserRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_type = db_user.get("user_type", current_user.get("user_type", ""))
        responsible_id = db_user.get("_id")
        if user_type == "admin":
            team_members = await user_repo.find_many(
                {
                    "user_type": "collaborator",
                    "is_deleted": {"$ne": True}
                },
                skip=0,
                limit=1000,
                sort=[("name", 1)]
            )
        else:
            team_members = await user_repo.find_by_responsible(responsible_id)
        
        formatted_members = []
        for member in team_members:
            formatted_members.append({
                "id": str(member.get("_id", "")),
                "name": member.get("name", "Unknown"),
                "email": member.get("email", ""),
            })
        
        return {"members": formatted_members}

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching team members: {str(err)}"
        )

@router.get("/pointage/entries/week/{week_start}")
async def get_pointage_entries_for_week(week_start: str, current_user: dict = RequireCollaborator):
    """
    Get all pointage entries for a specific week for the current collaborator.

    Args:
        week_start: Start date of the week in YYYY-MM-DD format (Monday of the week)
        current_user: Authenticated collaborator user

    Returns:
        Dictionary with entries list and week_start date
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")
        try:
            week_start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use YYYY-MM-DD"
            )
        cstr_semaine = get_cstr_semaine(week_start_date)
        user_id_str = str(user_id) if isinstance(user_id, ObjectId) else user_id
        query = {
            "user_id": {"$in": [user_id, user_id_str]},
            "entry_data.cstr_semaine": cstr_semaine,
            "is_deleted": {"$ne": True},
        }
        entries = await pointage_repo.find_many(query, sort=[("entry_data.date_pointage", 1)])

        formatted_entries = []
        for entry in entries:
            entry_data = entry.get("entry_data", {})
            formatted_entries.append({
                "id": str(entry.get("_id", "")),
                "date_pointage": serialize_date(entry_data.get("date_pointage")),
                "clef_imputation": entry_data.get("clef_imputation", ""),
                "libelle": entry_data.get("libelle", ""),
                "fonction": entry_data.get("fonction", ""),
                "date_besoin": serialize_date(entry_data.get("date_besoin")),
                "heures_theoriques": entry_data.get("heures_theoriques", ""),
                "heures_passees": entry_data.get("heures_passees", ""),
                "commentaires": entry_data.get("commentaires", ""),
                "status": entry.get("status", "draft"),
                "submitted_at": entry.get("submitted_at"),
                "created_at": entry.get("created_at"),
                "updated_at": entry.get("updated_at"),
            })

        return {"entries": formatted_entries, "week_start": week_start}

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching pointage entries: {str(e)}"
        )

@router.post("/pointage/entries")
async def create_pointage_entry(entry_data: PointageEntryCreate, current_user: dict = RequireCollaborator):
    """
    Create a new pointage entry for the current collaborator.

    Args:
        entry_data: Pointage entry data (date_pointage, LC fields, hours, etc.)
        current_user: Authenticated collaborator user

    Returns:
        Dictionary with created entry ID, message, and status
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")
        try:
            date_pointage_obj = datetime.strptime(entry_data.date_pointage, "%Y-%m-%d").date()
            date_besoin_obj = datetime.strptime(entry_data.date_besoin, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use YYYY-MM-DD"
            )

        week_start = date_pointage_obj - timedelta(days=date_pointage_obj.weekday())
        cstr_semaine = get_cstr_semaine(week_start)
        pointage_entry_data = PointageEntryData(
            date_pointage=date_pointage_obj,
            cstr_semaine=cstr_semaine,
            clef_imputation=entry_data.clef_imputation,
            libelle=entry_data.libelle,
            fonction=entry_data.fonction,
            date_besoin=date_besoin_obj,
            heures_theoriques=entry_data.heures_theoriques,
            heures_passees=entry_data.heures_passees,
            commentaires=entry_data.commentaires,
        )

        pointage_entry = PointageEntry(
            user_id=user_id,
            entry_data=pointage_entry_data,
            status="draft"
        )

        entry_id = await pointage_repo.create(pointage_entry)

        return {
            "id": str(entry_id),
            "message": "Pointage entry created successfully",
            "status": "draft"
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating pointage entry: {str(err)}"
        )

@router.put("/pointage/entries/{entry_id}")
async def update_pointage_entry(entry_id: str, entry_data: PointageEntryUpdate, current_user: dict = RequireCollaborator):
    """
    Update an existing pointage entry (only if status is draft).

    Args:
        entry_id: ID of the pointage entry to update
        entry_data: Updated pointage entry data
        current_user: Authenticated collaborator user

    Returns:
        Dictionary with entry ID, message, and updated status
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")
        if not ObjectId.is_valid(entry_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid entry ID"
            )

        entry_object_id = ObjectId(entry_id)
        existing_entry = await pointage_repo.find_by_id(entry_object_id)
        if not existing_entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pointage entry not found"
            )

        existing_user_id = existing_entry.get("user_id")
        existing_user_id_str = str(existing_user_id) if existing_user_id else None
        user_id_str = str(user_id) if user_id else None
        if existing_user_id_str != user_id_str:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only update your own entries"
            )

        if existing_entry.get("status") == "submitted":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot update submitted entry. It is locked."
            )

        try:
            date_besoin_obj = datetime.strptime(entry_data.date_besoin, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date_besoin format. Use YYYY-MM-DD"
            )

        existing_entry_data = existing_entry.get("entry_data", {})
        existing_date_pointage = existing_entry_data.get("date_pointage")
        if isinstance(existing_date_pointage, str):
            date_pointage_obj = datetime.strptime(existing_date_pointage, "%Y-%m-%d").date()
        elif isinstance(existing_date_pointage, date):
            date_pointage_obj = existing_date_pointage
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date_pointage in existing entry"
            )

        existing_date_besoin = None
        if not date_besoin_obj and existing_entry_data.get("date_besoin"):
            existing_date_besoin_str = existing_entry_data.get("date_besoin")
            if isinstance(existing_date_besoin_str, str):
                existing_date_besoin = datetime.strptime(existing_date_besoin_str, "%Y-%m-%d").date()
            elif isinstance(existing_date_besoin_str, date):
                existing_date_besoin = existing_date_besoin_str

        updated_entry_data = PointageEntryData(
            date_pointage=date_pointage_obj,
            cstr_semaine=existing_entry_data.get("cstr_semaine"),
            clef_imputation=entry_data.clef_imputation if entry_data.clef_imputation is not None else existing_entry_data.get("clef_imputation"),
            libelle=entry_data.libelle if entry_data.libelle is not None else existing_entry_data.get("libelle"),
            fonction=entry_data.fonction if entry_data.fonction is not None else existing_entry_data.get("fonction"),
            date_besoin=date_besoin_obj if date_besoin_obj else existing_date_besoin,
            heures_theoriques=entry_data.heures_theoriques if entry_data.heures_theoriques is not None else existing_entry_data.get("heures_theoriques"),
            heures_passees=entry_data.heures_passees if entry_data.heures_passees is not None else existing_entry_data.get("heures_passees"),
            commentaires=entry_data.commentaires if entry_data.commentaires is not None else existing_entry_data.get("commentaires"),
        )

        existing_status = existing_entry.get("status", "draft")
        updated_entry = PointageEntry(
            user_id=user_id,
            entry_data=updated_entry_data,
            status=existing_status
        )

        await pointage_repo.update(entry_object_id, updated_entry, current_user.get("email", "system"))
        return {
            "id": entry_id,
            "message": "Pointage entry updated successfully",
            "status": existing_status
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating pointage entry: {str(err)}"
        )

@router.post("/pointage/entries/{entry_id}/submit")
async def submit_pointage_entry(entry_id: str, current_user: dict = RequireCollaborator):
    """
    Submit a pointage entry (locks it for validation).

    Args:
        entry_id: ID of the pointage entry to submit
        current_user: Authenticated collaborator user

    Returns:
        Dictionary with entry ID, message, and submitted status
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")
        if not ObjectId.is_valid(entry_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid entry ID"
            )

        entry_object_id = ObjectId(entry_id)
        existing_entry = await pointage_repo.find_by_id(entry_object_id)
        if not existing_entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pointage entry not found"
            )

        existing_user_id = existing_entry.get("user_id")
        existing_user_id_str = str(existing_user_id) if existing_user_id else None
        user_id_str = str(user_id) if user_id else None
        if existing_user_id_str != user_id_str:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only submit your own entries"
            )

        if existing_entry.get("status") == "submitted":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Entry is already submitted"
            )

        await pointage_repo.submit(entry_object_id)
        return {
            "id": entry_id,
            "message": "Pointage entry submitted successfully",
            "status": "submitted"
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error submitting pointage entry: {str(e)}"
        )

@router.put("/pointage/entries/{entry_id}/status")
async def update_pointage_entry_status(
    entry_id: str, 
    status_data: dict = Body(...),
    current_user: dict = RequireAdminOrResponsible()):
    """
    Update the status of a pointage entry (for responsible/admin users only).
    
    Args:
        entry_id: ID of the pointage entry to update
        status_data: Dictionary with "status" field (draft, submitted, validated, rejected)
        current_user: Authenticated user (admin or responsible)
    
    Returns:
        Dictionary with entry ID, message, and updated status
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        
        if not ObjectId.is_valid(entry_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid entry ID"
            )
        
        new_status = status_data.get("status")
        if new_status not in ["draft", "submitted"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid status. Must be either 'draft' or 'submitted'"
            )
        
        entry_object_id = ObjectId(entry_id)
        existing_entry = await pointage_repo.find_by_id(entry_object_id)
        if not existing_entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pointage entry not found"
            )
        
        # Update status
        update_dict = {
            "status": new_status,
            "updated_at": datetime.utcnow(),
        }
        
        if new_status == "submitted":
            update_dict["submitted_at"] = datetime.utcnow()

        elif existing_entry.get("status") == "submitted" and new_status == "draft":
            update_dict["submitted_at"] = None
        
        await pointage_repo.collection.update_one(
            {"_id": entry_object_id},
            {"$set": update_dict}
        )
        
        return {
            "id": entry_id,
            "message": f"Entry status updated to {new_status}",
            "status": new_status
        }
    
    except HTTPException:
        raise
    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating entry status: {str(err)}"
        )

@router.delete("/pointage/entries/{entry_id}")
async def delete_pointage_entry(entry_id: str, current_user: dict = RequireCollaborator):
    """
    Delete a pointage entry (soft delete - marks as deleted).

    Args:
        entry_id: ID of the pointage entry to delete
        current_user: Authenticated collaborator user

    Returns:
        Dictionary with entry ID and message
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")
        if not ObjectId.is_valid(entry_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid entry ID"
            )

        entry_object_id = ObjectId(entry_id)
        existing_entry = await pointage_repo.find_by_id(entry_object_id)
        if not existing_entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pointage entry not found"
            )

        if existing_entry.get("user_id") != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only delete your own entries"
            )

        if existing_entry.get("status") == "submitted":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot delete submitted entry. It is locked."
            )

        await pointage_repo.mark_as_deleted(entry_object_id, str(user_id))
        return {
            "id": entry_id,
            "message": "Pointage entry deleted successfully"
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting pointage entry: {str(e)}"
        )

@router.post("/pointage/modification-requests")
async def create_modification_request(request_data: ModificationRequestCreate, current_user: dict = RequireCollaborator):
    """
    Create a modification request for a submitted entry.

    Args:
        request_data: Modification request data (entry_id, requested_data, comment)
        current_user: Authenticated collaborator user

    Returns:
        Dictionary with request ID and message
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        modification_repo = ModificationRequestRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")

        if not ObjectId.is_valid(request_data.entry_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid entry ID"
            )

        entry_object_id = ObjectId(request_data.entry_id)
        existing_entry = await pointage_repo.find_by_id(entry_object_id)
        if not existing_entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pointage entry not found"
            )

        existing_user_id = existing_entry.get("user_id")
        existing_user_id_str = str(existing_user_id) if existing_user_id else None
        user_id_str = str(user_id) if user_id else None
        if existing_user_id_str != user_id_str:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only request modification for your own entries"
            )
        
        if existing_entry.get("status") != "submitted":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Can only request modification for submitted entries"
            )

        existing_requests = await modification_repo.find_many(
            {
                "entry_id": entry_object_id,
                "status": "pending",
                "is_deleted": {"$ne": True}
            }
        )

        if existing_requests:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A pending modification request already exists for this entry"
            )

        modification_request = ModificationRequest(
            entry_id=entry_object_id,
            user_id=user_id,
            requested_data=request_data.requested_data.model_dump(),
            comment=request_data.comment,
            status="pending"
        )

        request_id = await modification_repo.create(modification_request)
        return {
            "id": str(request_id),
            "message": "Modification request created successfully"
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating modification request: {str(err)}"
        )

@router.get("/pointage/modification-requests")
async def get_modification_requests(
    current_user: dict = RequireAdminOrResponsible(), 
    skip: int = 0, 
    limit: int = 100,
    status: Optional[str] = None):
    """
    Get modification requests for a responsible's team (or all for admin).

    Args:
        current_user: Authenticated user (admin or responsible)
        skip: Number of requests to skip
        limit: Maximum number of requests to return
        status: Optional status filter ("pending", "approved", "rejected"). If None, returns all requests.

    Returns:
        Dictionary with requests list and metadata
    """
    try:
        user_repo = UserRepository()
        modification_repo = ModificationRequestRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)

        user_type = db_user.get("user_type", current_user.get("user_type", ""))
        responsible_id = db_user.get("_id")
        if user_type == "admin":
            query = {"is_deleted": {"$ne": True}}
            if status:
                query["status"] = status
            requests = await modification_repo.find_many(
                query,
                skip=skip,
                limit=limit,
                sort=[("created_at", -1)]
            )
        else:
            requests = await modification_repo.find_by_team(
                responsible_id,
                skip=skip,
                limit=limit
            )
            if status:
                requests = [r for r in requests if r.get("status") == status]
        
        formatted_requests = []
        for req in requests:
            entry_id = req.get("entry_id")
            user_id = req.get("user_id")
            entry = None
            if entry_id:
                if isinstance(entry_id, str) and ObjectId.is_valid(entry_id):
                    entry = await pointage_repo.find_by_id(ObjectId(entry_id))
                elif isinstance(entry_id, ObjectId):
                    entry = await pointage_repo.find_by_id(entry_id)

            user_info = {"name": "Unknown", "email": ""}
            if user_id:
                try:
                    if isinstance(user_id, str) and ObjectId.is_valid(user_id):
                        user_obj = await user_repo.find_by_id(ObjectId(user_id))
                    elif isinstance(user_id, ObjectId):
                        user_obj = await user_repo.find_by_id(user_id)
                    else:
                        user_obj = None
                    
                    if user_obj:
                        user_info = {
                            "name": user_obj.get("name", "Unknown"),
                            "email": user_obj.get("email", "")
                        }
                except Exception:
                    pass

            entry_data = entry.get("entry_data", {}) if entry else {}
            formatted_requests.append({
                "id": str(req.get("_id", "")),
                "entry_id": str(entry_id) if entry_id else "",
                "user_id": str(user_id) if user_id else "",
                "user_name": user_info.get("name", "Unknown"),
                "user_email": user_info.get("email", ""),
                "requested_data": req.get("requested_data", {}),
                "current_data": {
                    "clef_imputation": entry_data.get("clef_imputation", ""),
                    "libelle": entry_data.get("libelle", ""),
                    "fonction": entry_data.get("fonction", ""),
                    "date_besoin": serialize_date(entry_data.get("date_besoin")),
                    "heures_theoriques": entry_data.get("heures_theoriques", ""),
                    "heures_passees": entry_data.get("heures_passees", ""),
                    "commentaires": entry_data.get("commentaires", ""),
                },
                "date_pointage": serialize_date(entry_data.get("date_pointage")) if entry else "",
                "comment": req.get("comment"),
                "status": req.get("status", "pending"),
                "created_at": req.get("created_at"),
                "reviewed_at": req.get("reviewed_at"),
                "reviewed_by": req.get("reviewed_by"),
                "review_comment": req.get("review_comment"),
            })

        return {
            "requests": formatted_requests,
            "total": len(formatted_requests),
            "skip": skip,
            "limit": limit
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching modification requests: {str(err)}"
        )

@router.post("/pointage/modification-requests/{request_id}/review")
async def review_modification_request(request_id: str, review_data: ModificationRequestReview, current_user: dict = RequireAdminOrResponsible()):
    """
    Review (approve or reject) a modification request.

    Args:
        request_id: ID of the modification request
        review_data: Review decision (status: "approved" or "rejected", optional review_comment)
        current_user: Authenticated user (admin or responsible)

    Returns:
        Dictionary with request ID and message
    """
    try:
        user_repo = UserRepository()
        pointage_repo = PointageEntryRepository()
        modification_repo = ModificationRequestRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        if review_data.status not in ["approved", "rejected"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Status must be 'approved' or 'rejected'"
            )

        if not ObjectId.is_valid(request_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid request ID"
            )

        request_object_id = ObjectId(request_id)
        existing_request = await modification_repo.find_by_id(request_object_id)
        if not existing_request:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Modification request not found"
            )

        if existing_request.get("status") != "pending":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Request has already been reviewed"
            )

        entry_id = existing_request.get("entry_id")
        if not entry_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid entry ID in request"
            )

        entry_object_id = entry_id if isinstance(entry_id, ObjectId) else ObjectId(entry_id)
        entry = await pointage_repo.find_by_id(entry_object_id)
        if not entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pointage entry not found"
            )

        update_dict = {
            "status": review_data.status,
            "reviewed_at": datetime.utcnow(),
            "reviewed_by": db_user.get("email", current_user.get("email", "system")),
        }
        if review_data.review_comment:
            update_dict["review_comment"] = review_data.review_comment

        await modification_repo.collection.update_one(
            {"_id": request_object_id},
            {"$set": update_dict}
        )

        if review_data.status == "approved":
            requested_data = existing_request.get("requested_data", {})
            existing_entry_data = entry.get("entry_data", {})
            date_besoin_str = requested_data.get("date_besoin")
            if date_besoin_str:
                try:
                    date_besoin_obj = datetime.strptime(date_besoin_str, "%Y-%m-%d").date()
                except ValueError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid date_besoin format in requested data"
                    )
            else:
                existing_date_besoin = existing_entry_data.get("date_besoin")
                if isinstance(existing_date_besoin, str):
                    date_besoin_obj = datetime.strptime(existing_date_besoin, "%Y-%m-%d").date()
                elif isinstance(existing_date_besoin, date):
                    date_besoin_obj = existing_date_besoin
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid date_besoin in existing entry"
                    )

            existing_date_pointage = existing_entry_data.get("date_pointage")
            if isinstance(existing_date_pointage, str):
                date_pointage_obj = datetime.strptime(existing_date_pointage, "%Y-%m-%d").date()
            elif isinstance(existing_date_pointage, date):
                date_pointage_obj = existing_date_pointage
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid date_pointage in existing entry"
                )

            updated_entry_data = PointageEntryData(
                date_pointage=date_pointage_obj,
                cstr_semaine=existing_entry_data.get("cstr_semaine"),
                clef_imputation=requested_data.get("clef_imputation", existing_entry_data.get("clef_imputation", "")),
                libelle=requested_data.get("libelle", existing_entry_data.get("libelle", "")),
                fonction=requested_data.get("fonction", existing_entry_data.get("fonction", "")),
                date_besoin=date_besoin_obj,
                heures_theoriques=requested_data.get("heures_theoriques", existing_entry_data.get("heures_theoriques", "")),
                heures_passees=requested_data.get("heures_passees", existing_entry_data.get("heures_passees", "")),
                commentaires=requested_data.get("commentaires", existing_entry_data.get("commentaires")),
            )

            updated_entry = PointageEntry(
                user_id=entry.get("user_id"),
                entry_data=updated_entry_data,
                status="draft"
            )
            await pointage_repo.update(entry_object_id, updated_entry, db_user.get("email", "system"))

        return {
            "id": request_id,
            "message": f"Modification request {review_data.status} successfully"
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error reviewing modification request: {str(err)}"
        )

@router.get("/pointage/modification-requests/my-requests")
async def get_my_modification_requests(current_user: dict = RequireCollaborator, skip: int = 0, limit: int = 100):
    """
    Get modification requests for the current collaborator.

    Args:
        current_user: Authenticated collaborator user
        skip: Number of requests to skip
        limit: Maximum number of requests to return

    Returns:
        Dictionary with requests list and metadata
    """
    try:
        user_repo = UserRepository()
        modification_repo = ModificationRequestRepository()
        pointage_repo = PointageEntryRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_id = db_user.get("_id")

        requests = await modification_repo.find_by_user(
            user_id,
            skip=skip,
            limit=limit
        )

        # Include all requests (pending, approved, rejected)
        formatted_requests = []
        for req in requests:
            entry_id = req.get("entry_id")
            entry = None
            if entry_id:
                if isinstance(entry_id, str) and ObjectId.is_valid(entry_id):
                    entry = await pointage_repo.find_by_id(ObjectId(entry_id))
                elif isinstance(entry_id, ObjectId):
                    entry = await pointage_repo.find_by_id(entry_id)

            entry_data = entry.get("entry_data", {}) if entry else {}
            requested_data = req.get("requested_data", {})
            formatted_requests.append({
                "id": str(req.get("_id", "")),
                "entry_id": str(entry_id) if entry_id else "",
                "status": req.get("status", ""),
                "review_comment": req.get("review_comment"),
                "reviewed_at": req.get("reviewed_at"),
                "created_at": req.get("created_at"),
                "date_pointage": serialize_date(entry_data.get("date_pointage")) if entry else "",
                "cstr_semaine": entry_data.get("cstr_semaine") if entry else "",
                "requested_data": requested_data,
                "current_data": {
                    "clef_imputation": entry_data.get("clef_imputation", ""),
                    "libelle": entry_data.get("libelle", ""),
                    "fonction": entry_data.get("fonction", ""),
                    "date_besoin": serialize_date(entry_data.get("date_besoin")),
                    "heures_theoriques": entry_data.get("heures_theoriques", ""),
                    "heures_passees": entry_data.get("heures_passees", ""),
                    "commentaires": entry_data.get("commentaires", ""),
                },
                "comment": req.get("comment"),
            })

        return {
            "requests": formatted_requests,
            "total": len(formatted_requests),
            "skip": skip,
            "limit": limit
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching modification requests: {str(err)}"
        )


@router.get("/conditional-lists/default/all-items")
async def get_all_lc_items(current_user: dict = RequireAdminOrResponsible()):
    """
    Get all items from the active LC (Liste Conditionnelle) for admin editing.
    Returns all items including inactive ones.
    """
    try:
        repo = ConditionalListRepository()
        active_lc_name = await get_active_lc_name()
        active_lc = await repo.find_by_name(active_lc_name)
        if not active_lc:
            return {"items": []}

        items = active_lc.get("items", [])
        formatted_items = []
        for idx, item in enumerate(items):
            formatted_items.append({
                "index": idx,
                "clef_imputation": item.get("clef_imputation", ""),
                "libelle": item.get("libelle", ""),
                "fonction": item.get("fonction", ""),
                "is_active": item.get("is_active", True),
            })

        return {"items": formatted_items}

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching LC items: {str(e)}"
        )


@router.put("/conditional-lists/default/items/update")
async def update_lc_item(update_data: LCItemUpdate, current_user: dict = RequireAdminOrResponsible()):
    """
    Update a single cell in an LC item.
    Each cell (clef_imputation, libelle, fonction) can be updated independently.
    """
    try:
        repo = ConditionalListRepository()
        active_lc_name = await get_active_lc_name()
        active_lc = await repo.find_by_name(active_lc_name)
        if not active_lc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Active LC '{active_lc_name}' not found"
            )

        items = active_lc.get("items", [])
        if update_data.item_index < 0 or update_data.item_index >= len(items):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid item index: {update_data.item_index}"
            )

        if update_data.field not in ["clef_imputation", "libelle", "fonction"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid field: {update_data.field}. Must be one of: clef_imputation, libelle, fonction"
            )

        update_path = f"items.{update_data.item_index}.{update_data.field}"
        update_dict = {
            update_path: update_data.value,
            "updated_at": datetime.utcnow(),
            "updated_by": current_user.get("email", "system"),
        }

        if update_data.is_active is not None:
            update_dict[f"items.{update_data.item_index}.is_active"] = update_data.is_active

        result = await repo.collection.update_one(
            {"_id": active_lc["_id"]},
            {"$set": update_dict}
        )

        if result.modified_count == 0:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to update LC item"
            )

        return {
            "message": "LC item updated successfully",
            "item_index": update_data.item_index,
            "field": update_data.field,
            "value": update_data.value
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating LC item: {str(e)}"
        )

@router.get("/conditional-lists/all")
async def get_all_conditional_lists(current_user: dict = RequireAdminOrResponsible()):
    """
    Get all conditional lists (names only) for admin selection.
    Returns list of all conditional list names.
    """
    try:
        repo = ConditionalListRepository()
        lists = await repo.find_active_lists(skip=0, limit=1000)
        formatted_lists = []
        for lc in lists:
            if lc.get("name") != "_SYSTEM_ACTIVE_LC":
                formatted_lists.append({
                    "name": lc.get("name", ""),
                    "description": lc.get("description", ""),
                })

        return {"lists": formatted_lists}
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching conditional lists: {str(e)}"
        )


@router.get("/conditional-lists/active")
async def get_active_conditional_list(current_user: dict = RequireAdminOrResponsible()):
    """
    Get the name of the currently active conditional list.
    """
    try:
        active_name = await get_active_lc_name()
        return {"active_lc_name": active_name}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching active LC: {str(e)}"
        )

@router.put("/conditional-lists/active")
async def set_active_conditional_list(update_data: ActiveLCUpdate, current_user: dict = RequireAdminOrResponsible()):
    """
    Set the active conditional list that will be used system-wide.
    """
    try:
        success = await set_active_lc_name(update_data.lc_name)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Conditional list '{update_data.lc_name}' not found"
            )

        return {
            "message": f"Active conditional list set to '{update_data.lc_name}'",
            "active_lc_name": update_data.lc_name
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error setting active LC: {str(e)}"
        )

@router.post("/conditional-lists")
async def create_conditional_list(list_data: ConditionalListCreate, current_user: dict = RequireAdminOrResponsible()):
    """
    Create a new conditional list with items.
    """
    try:
        repo = ConditionalListRepository()
        db_user = await get_db_user_from_current(current_user, UserRepository())
        
        # Check if name already exists
        existing = await repo.find_by_name(list_data.name)
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Conditional list with name '{list_data.name}' already exists"
            )
        
        # Convert items to ConditionalListItem
        items = [
            ConditionalListItem(
                clef_imputation=item.clef_imputation,
                libelle=item.libelle,
                fonction=item.fonction,
                is_active=item.is_active
            )
            for item in list_data.items
        ]
        
        conditional_list = ConditionalList(
            name=list_data.name,
            description=list_data.description,
            items=items,
            created_by=db_user.get("email", current_user.get("email", "system")),
            updated_by=db_user.get("email", current_user.get("email", "system"))
        )
        
        lc_id = await repo.create(conditional_list)
        
        return {
            "id": str(lc_id),
            "name": list_data.name,
            "message": f"Conditional list '{list_data.name}' created successfully with {len(items)} items"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating conditional list: {str(e)}"
        )


@router.post("/conditional-lists/merge")
async def merge_lc_items(merge_data: LCMergeRequest, current_user: dict = RequireAdminOrResponsible()):
    """
    Merge items into an existing conditional list, removing duplicates if specified.
    """
    try:
        repo = ConditionalListRepository()
        db_user = await get_db_user_from_current(current_user, UserRepository())
        active_lc_name = await get_active_lc_name()

        target_lc = await repo.find_by_name(merge_data.lc_name)
        if not target_lc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Conditional list '{merge_data.lc_name}' not found"
            )

        existing_items = target_lc.get("items", [])

        existing_clef_imputation = set()
        existing_libelle = set()
        existing_fonction = set()
        if merge_data.remove_duplicates:
            for item in existing_items:
                clef = item.get("clef_imputation", "")
                libelle = item.get("libelle", "")
                fonction = item.get("fonction", "")
                if clef:
                    existing_clef_imputation.add(clef)
                if libelle:
                    existing_libelle.add(libelle)
                if fonction:
                    existing_fonction.add(fonction)

        new_items = []
        duplicates_count = 0
        for item in merge_data.items:
            if merge_data.remove_duplicates:
                is_duplicate = False
                if item.clef_imputation and item.clef_imputation in existing_clef_imputation:
                    is_duplicate = True
                if item.libelle and item.libelle in existing_libelle:
                    is_duplicate = True
                if item.fonction and item.fonction in existing_fonction:
                    is_duplicate = True

                if is_duplicate:
                    duplicates_count += 1
                    continue

            if item.clef_imputation:
                existing_clef_imputation.add(item.clef_imputation)
            if item.libelle:
                existing_libelle.add(item.libelle)
            if item.fonction:
                existing_fonction.add(item.fonction)

            new_items.append({
                "clef_imputation": item.clef_imputation,
                "libelle": item.libelle,
                "fonction": item.fonction,
                "is_active": item.is_active
            })
        if new_items:
            for item in new_items:
                await repo.add_item(target_lc["_id"], item, db_user.get("email", current_user.get("email", "system")))

        return {
            "message": f"Merged {len(new_items)} new items into '{merge_data.lc_name}'",
            "added": len(new_items),
            "duplicates_skipped": duplicates_count,
            "total_items": len(existing_items) + len(new_items)
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error merging items: {str(e)}"
        )


@router.post("/conditional-lists/parse-excel")
async def parse_excel_file(
    file: UploadFile = File(...),
    current_user: dict = RequireAdminOrResponsible()
):
    """
    Parse an Excel file and extract LC items.
    Expected format:
    - Row 2: Headers (Clef d'imputation, Libellé, Fonction)
    - Row 3+: Data rows
    """
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be an Excel file (.xlsx or .xls)"
            )

        contents = await file.read()
        try:
            workbook = load_workbook(BytesIO(contents), data_only=True)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file format: {str(e)}"
            )

        if not workbook.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file has no sheets"
            )

        sheet = workbook[workbook.sheetnames[0]]
        if sheet.max_row < 2:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have at least 2 rows (header and data)"
            )

        headers = []
        for cell in sheet[2]:
            headers.append(str(cell.value or '').strip())

        header_map = {}
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if 'clef' in header_lower or 'imputation' in header_lower:
                header_map['clef_imputation'] = idx + 1
                continue

            if 'libellé' in header_lower or 'libelle' in header_lower:
                header_map['libelle'] = idx + 1
                continue

            if 'fonction' in header_lower:
                header_map['fonction'] = idx + 1
                continue

        if 'clef_imputation' not in header_map or 'libelle' not in header_map or 'fonction' not in header_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have columns: Clef d'imputation, Libellé, Fonction"
            )

        items = []
        for row_num in range(3, sheet.max_row + 1):
            clef_imputation = str(sheet.cell(row=row_num, column=header_map['clef_imputation']).value or '').strip()
            libelle = str(sheet.cell(row=row_num, column=header_map['libelle']).value or '').strip()
            fonction = str(sheet.cell(row=row_num, column=header_map['fonction']).value or '').strip()
            if not clef_imputation and not libelle and not fonction:
                continue

            items.append({
                "clef_imputation": clef_imputation if clef_imputation else "-",
                "libelle": libelle if libelle else "-",
                "fonction": fonction if fonction else "-",
                "is_active": True
            })

        if not items:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid data rows found in Excel file"
            )

        return {
            "items": items,
            "count": len(items)
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error parsing Excel file: {str(e)}"
        )


@router.get("/users/all")
async def get_all_users(current_user: dict = RequireAdminOrResponsible()):
    """
    Get all users (collaborators and responsibles) for admin.
    For responsible users, returns their team members only.
    """
    try:
        user_repo = UserRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        user_type = db_user.get("user_type", current_user.get("user_type", ""))

        if user_type == "admin":
            collaborators = await user_repo.find_many(
                {
                    "user_type": "collaborator",
                    "is_deleted": {"$ne": True}
                },
                skip=0,
                limit=1000,
                sort=[("name", 1)]
            )
            responsibles = await user_repo.find_responsibles(skip=0, limit=1000)
            all_users = collaborators + responsibles
        else:
            responsible_id = db_user.get("_id")
            all_users = await user_repo.find_by_responsible(responsible_id, skip=0, limit=1000)

        formatted_users = []
        for user in all_users:
            responsible_id = user.get("responsible_id")
            formatted_users.append({
                "id": str(user.get("_id", "")),
                "name": user.get("name", "Unknown"),
                "email": user.get("email", ""),
                "user_type": user.get("user_type", ""),
                "status": user.get("status", "active"),
                "responsible_id": str(responsible_id) if responsible_id else None,
            })

        return {"users": formatted_users}

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error fetching users: {str(err)}"
        )

@router.post("/users")
async def create_user(user_data: UserCreate, current_user: dict = RequireAdminOrResponsible()):
    """
    Create a new user (collaborator or responsible).
    """
    try:
        user_repo = UserRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        if user_data.user_type not in ["collaborator", "responsible"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="user_type must be 'collaborator' or 'responsible'"
            )

        if user_data.status not in ["active", "inactive"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="status must be 'active' or 'inactive'"
            )

        if user_data.email:
            existing = await user_repo.find_by_email(user_data.email)
            if existing:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"User with email {user_data.email} already exists"
                )

        responsible_id = None
        if user_data.responsible_id:
            if ObjectId.is_valid(user_data.responsible_id):
                responsible_id = ObjectId(user_data.responsible_id)
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid responsible_id format"
                )

        user = User(
            name=user_data.name,
            email=user_data.email,
            user_type=user_data.user_type,
            status=user_data.status,
            responsible_id=responsible_id,
            created_by=db_user.get("email", current_user.get("email", "system")),
            updated_by=db_user.get("email", current_user.get("email", "system"))
        )

        user_id = await user_repo.create(user)

        return {
            "id": str(user_id),
            "message": "User created successfully"
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating user: {str(err)}"
        )

@router.put("/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdate, current_user: dict = RequireAdminOrResponsible()):
    """
    Update an existing user.
    """
    try:
        user_repo = UserRepository()
        db_user = await get_db_user_from_current(current_user, user_repo)
        if not ObjectId.is_valid(user_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid user ID"
            )

        user_object_id = ObjectId(user_id)
        existing_user = await user_repo.find_by_id(user_object_id)
        if not existing_user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        update_dict = {}
        if user_data.name is not None:
            update_dict["name"] = user_data.name
        if user_data.email is not None:
            existing_by_email = await user_repo.find_by_email(user_data.email)
            if existing_by_email and str(existing_by_email.get("_id")) != user_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Email {user_data.email} is already taken by another user"
                )
            update_dict["email"] = user_data.email

        if user_data.user_type is not None:
            if user_data.user_type not in ["collaborator", "responsible", "admin"]:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="user_type must be 'collaborator', 'responsible', or 'admin'"
                )
            update_dict["user_type"] = user_data.user_type

        if user_data.status is not None:
            if user_data.status not in ["active", "inactive"]:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="status must be 'active' or 'inactive'"
                )
            update_dict["status"] = user_data.status

        if user_data.responsible_id is not None:
            if user_data.responsible_id:
                if ObjectId.is_valid(user_data.responsible_id):
                    update_dict["responsible_id"] = ObjectId(user_data.responsible_id)
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid responsible_id format"
                    )
            else:
                update_dict["responsible_id"] = None

        updated_user_data = {**existing_user, **update_dict}
        updated_user = User(**updated_user_data)
        await user_repo.update(user_object_id, updated_user, db_user.get("email", current_user.get("email", "system")))
        return {
            "id": user_id,
            "message": "User updated successfully"
        }

    except Exception as err:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating user: {str(err)}"
        )
